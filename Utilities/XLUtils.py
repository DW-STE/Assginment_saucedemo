import openpyxl

class XLUtils:


        def readData(file,sheetName,rownum,columnnum):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook[sheetName]
            return sheet.cell(row=rownum,column=columnnum).value

        def writeData(file,sheetName,rownum,columnnum,data):
            workbook = openpyxl.load_workbook(file)
            sheet = workbook[sheetName]
            sheet.cell(row=rownum,column= columnnum).value= data
            workbook.save(file)




